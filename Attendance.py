import serial
import time
from datetime import  date
from openpyxl import load_workbook
#from openpyxl.utils.cell import get_column_letter
import openpyxl
# Open the serial port
#ser = serial.Serial('COM4', 9600)  # Update 'COM3' to match your Arduino's serial port

# Load the Excel file
workbook = load_workbook(r"C:\Users\Bharani\OneDrive\Documents\attendance_system.xlsx")
worksheet = workbook.active
#start_date=date(2024,4,18)
#num_days=31

# Function to mark attendance
def next_column(sheet):
    maxcol=sheet.max_column
    return openpyxl.utils.get_column_letter(maxcol + 1)
   
def mark_attendance(id_number):
   
        #worksheet.cell(row=id_number+1, column=dat).value = "/"
        worksheet[nxt+str(id_number+1)]='/'
        workbook.save(r"C:\Users\Bharani\OneDrive\Documents\attendance_system.xlsx")
        rollnumber=worksheet['B'+str(id_number+1)].value
        print('Attendance marked for {} '.format(rollnumber))
   
        return
nxt=str(next_column(worksheet))

current_date=date.today()
head=current_date.strftime('%d-%m-%Y')
#worksheet.cell(row=1,column=,value=head)
worksheet[nxt+"1"]=head
   
start_time = time.time()
ser = serial.Serial('COM4', 9600)

while True:
       
    if ser.in_waiting > 0:
        id_number = int(ser.readline().decode().strip())
        mark_attendance(id_number)

       
        elapsed_time = time.time() - start_time
        if elapsed_time >= 30:  
            print("2 minutes have elapsed. Attendance process is over")
            break  # Exit the loop
ser.close()

